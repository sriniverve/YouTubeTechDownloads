from usr.libs.youtube_downloads import YouTubeDownloads
from openpyxl import load_workbook
from openpyxl import Workbook
from usr.libs.format_excel_sheet import cellFormat as cf



API_KEY = "AIzaSyCJ2Bt5jCjYGPovjRq8HPRFimw2xDbT8cM"
API_KEY1 = "AIzaSyDbwi6sP_pKSrbk4QxkwoFPwaUW8cdgSHM"
#API_KEY2 =
API_KEY3 = "AIzaSyDAMLKs-D1pcf2Pw6aZFe3tJjdT1E17dzM"
filename = "youtube_tech_download_links.xlsx"
domain = ["802.11ac", "WiFi6", "Multicast", "WiFi Security", "Python", "IPSec", "AWS", "GCP", "IGMP", "Ansible", "Jenkins", "Advanced Python"]
#domain = ["802.11ac", "WiFi6"]


for item in domain:
    try:
        wb = load_workbook(filename=filename)
        sheet = wb.create_sheet()
    except:
        wb = Workbook()
        sheet = wb.active

    sheet.column_dimensions['A'].width = 100
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 30
    sheet['A1'] = "Title"
    sheet['B1'] = "Video Link"
    sheet['C1'] = "Views"
    cf(sheet['A1'], 'D3D3D3', 15, True, '000000', True)
    cf(sheet['B1'], 'D3D3D3', 15, True, '000000', True)
    cf(sheet['C1'], 'D3D3D3', 15, True, '000000', True)
    sheet.title = item

    yt = YouTubeDownloads(API_KEY3, item+" tutorial")
    video_data = yt.get_video_details()
    if video_data is None:
        exit(0)
    else:
        for list_items in range(2, len(video_data)):
            dict_item = video_data[list_items]
            sheet.cell(row=list_items, column=1).value = dict_item["Title"]
            sheet.cell(row=list_items, column=2).value = dict_item["Video Link"]
            sheet.cell(row=list_items, column=3).value = dict_item["Views"]
    wb.save(filename)